from ultralytics import YOLO
import os
from openpyxl import load_workbook 

def write_to_excel(dict, excel_path):
    wb = load_workbook(excel_path)
    ws = wb.active

    for key, values in dict.items():
        for r in range(3, len(dict)+3):
            if key == ws.cell(row=r, column=1).value:
                ws.cell(row=r, column=2).value = values

    wb.save(excel_path)

def tensor_to_list(tensor):
    list_array = tensor.tolist()
    for list_only in list_array:
        for i, _ in enumerate(list_only):
            list_only[i] = int(_)/1
        list_only.append(1.0)
    return str(list_array)[1:-1]

def main(results, excel_path):
    dict = {}
    for _ in results:
        dict[os.path.basename(_.path)] = tensor_to_list(_.boxes.xyxy)
    
    write_to_excel(dict, excel_path)
    

if __name__ == '__main__':
    model = YOLO(r'C:\\Users\\iDo\\Desktop\\预处理前\\weights\\best.pt')
    folder_path = r'C:\\Users\\iDo\\Desktop\\3_Test\\Figures'
    excel_path = r'C:\\Users\\iDo\\Desktop\\3_Test\\Test_results.xlsx'

    results = model(folder_path, save=True)
    main(results, excel_path)